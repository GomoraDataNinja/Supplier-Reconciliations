import streamlit as st
import pandas as pd
import numpy as np
import re
import tempfile
from datetime import datetime, date
from dateutil.parser import parse as dt_parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy as pycopy

st.set_page_config(page_title="Recon App", layout="wide")


def to_str(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def to_num(x):
    if pd.isna(x):
        return np.nan
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = to_str(x).replace(",", "")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan

def to_date(x, dayfirst=False):
    if isinstance(x, (pd.Timestamp, np.datetime64)):
        return pd.to_datetime(x, errors="coerce")
    s = to_str(x)
    if not s:
        return pd.NaT
    try:
        return pd.to_datetime(dt_parse(s, fuzzy=True, dayfirst=dayfirst))
    except Exception:
        return pd.NaT

def round2(x):
    if pd.isna(x):
        return np.nan
    return float(round(float(x), 2))

def round0(x):
    if pd.isna(x):
        return np.nan
    return float(round(float(x), 0))

def date_diff_days(d1, d2):
    if pd.isna(d1) or pd.isna(d2):
        return 999999
    return abs((pd.to_datetime(d1).date() - pd.to_datetime(d2).date()).days)

def normalize_invoice(s: str) -> str:
    s = (s or "").upper().strip()
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def normalize_text(s: str) -> str:
    s = (s or "").upper()
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.replace("/", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def row_non_empty_count(row):
    return int(row.notna().sum())

def looks_like_header(cell):
    if pd.isna(cell):
        return False
    s = str(cell).strip().lower()
    if not s:
        return False
    tokens = [
        "date", "invoice", "inv", "debit", "credit", "amount",
        "ref", "reference", "details", "description",
        "doc", "posting", "external", "type", "balance"
    ]
    return any(t in s for t in tokens)

def detect_best_table_in_sheet(raw: pd.DataFrame, max_scan_rows: int = 80, sheet_name: str = ""):
    if raw is None or raw.empty:
        return None, None

    scan_rows = min(max_scan_rows, len(raw))
    best_score = -1
    best_df = None
    best_header_row = None

    for r in range(scan_rows):
        header_row = raw.iloc[r]
        non_empty = row_non_empty_count(header_row)
        if non_empty < 3:
            continue

        header_hits = sum(looks_like_header(x) for x in header_row.values)
        data_block = raw.iloc[r + 1: r + 1 + 30].copy()
        if data_block.empty:
            continue

        block_counts = [row_non_empty_count(data_block.iloc[i]) for i in range(min(len(data_block), 10))]
        consistency = float(np.mean(block_counts)) if block_counts else 0.0

        score = header_hits * 6 + non_empty + consistency

        if score > best_score:
            cols = [str(x).strip() if not pd.isna(x) else "" for x in header_row.values]
            df = raw.iloc[r + 1:].copy()
            df.columns = cols
            df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
            df = df.dropna(how="all").reset_index(drop=True)

            best_score = score
            best_df = df
            best_header_row = r

    if best_df is None:
        return None, None

    meta = {
        "sheet_name": sheet_name,
        "header_row": int(best_header_row),
        "score": float(best_score),
        "rows": int(len(best_df)),
        "cols": int(len(best_df.columns)),
    }
    return best_df, meta


def pick_col_by_hint(df: pd.DataFrame, hint: str):
    hint = (hint or "").strip().lower()
    if not hint:
        return ""
    cols = list(df.columns)
    hits = [c for c in cols if hint in str(c).lower()]
    if len(hits) == 1:
        return hits[0]
    if len(hits) > 1:
        lengths = [(c, len(str(c))) for c in hits]
        lengths.sort(key=lambda x: x[1])
        return lengths[0][0]
    return ""

def _is_date_like(x):
    try:
        dt_parse(str(x), fuzzy=True)
        return True
    except Exception:
        return False

def infer_col_by_type(df: pd.DataFrame, role: str):
    best_col = ""
    best_score = -1e9

    for c in df.columns:
        s = df[c].map(to_str)
        n = df[c].map(to_num)
        d = s.map(lambda x: True if x and _is_date_like(x) else False)

        pct_date = float(d.mean())
        pct_num = float(n.notna().mean())
        avg_len = float(s.map(len).mean())
        sparsity = float(df[c].isna().mean())

        header = str(c).lower()

        score = 0.0
        if role == "date":
            score = pct_date * 10 - sparsity * 2
            if "date" in header or "posting" in header:
                score += 3
        elif role == "amount":
            score = pct_num * 10 - sparsity * 1
            if "amount" in header or "amt" in header:
                score += 3
            if "debit" in header or "credit" in header:
                score -= 2
        elif role == "invoice":
            score = avg_len * 0.4 + float(s.nunique(dropna=True) / max(1, len(s))) * 2
            if "invoice" in header or "inv" in header:
                score += 5
            if "doc" in header and "invoice" not in header:
                score -= 1
        elif role == "external_doc":
            score = avg_len * 0.5 + float(s.nunique(dropna=True) / max(1, len(s))) * 1.5
            if "external" in header:
                score += 6
            if "doc" in header:
                score += 2
        elif role == "description":
            score = avg_len * 0.8
            if "desc" in header or "details" in header:
                score += 3

        if score > best_score:
            best_score = score
            best_col = c

    return best_col


AE_TOKEN = re.compile(r"\bAE[A-Z]{0,3}\d{4,}\b", re.IGNORECASE)

def extract_ae_candidates(text: str):
    t = normalize_text(text)
    found = AE_TOKEN.findall(t)
    return [normalize_invoice(x) for x in found]

def ledger_invoice_key(external_doc: str, supplier_invoice_set: set):
    candidates = extract_ae_candidates(external_doc)
    candidates = [c for c in candidates if c in supplier_invoice_set]
    if not candidates:
        return ""
    t = normalize_text(external_doc)
    last_pos = -1
    winner = candidates[0]
    for c in candidates:
        pos = t.rfind(c)
        if pos > last_pos:
            last_pos = pos
            winner = c
    return winner


DOCID_TOKEN = re.compile(r"(HREINV|HRECRN)\s*0*([0-9]+)", re.IGNORECASE)

def extract_docid(x):
    if pd.isna(x):
        return ""
    s = str(x).upper()
    m = DOCID_TOKEN.search(s)
    if not m:
        return ""
    prefix = m.group(1).upper()
    num = m.group(2)
    try:
        return prefix + str(int(num))
    except Exception:
        return prefix + num.lstrip("0")


def classify_ledger_txn(doc_type: str, external_doc: str, desc: str):
    text = f"{doc_type} {external_doc} {desc}".upper()
    if "CREDIT" in text and "MEMO" in text:
        return "credit"
    if text.startswith("CASJ") or "PAYMENT" in text or "RECEIPT" in text or "BANK" in text or "EFT" in text or "RTGS" in text:
        return "payment"
    return "invoice"


def normalize_supplier_sheet(df: pd.DataFrame, sheet_name: str, invoice_hint: str, date_hint: str, amount_hint: str, desc_hint: str):
    inv_col = pick_col_by_hint(df, invoice_hint) or infer_col_by_type(df, "invoice")
    date_col = pick_col_by_hint(df, date_hint) or infer_col_by_type(df, "date")
    amt_col = pick_col_by_hint(df, amount_hint) or infer_col_by_type(df, "amount")
    desc_col = pick_col_by_hint(df, desc_hint) or infer_col_by_type(df, "description")

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=False)) if date_col in df.columns else pd.NaT
    out["invoice_no_raw"] = df[inv_col].map(to_str) if inv_col in df.columns else ""
    out["invoice_key"] = out["invoice_no_raw"].map(normalize_invoice)
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""
    out["amount_signed"] = df[amt_col].map(to_num) if amt_col in df.columns else np.nan

    out["amt_r2"] = out["amount_signed"].map(round2)
    out["amt_r0"] = out["amount_signed"].map(round0)

    out["sheet_name"] = sheet_name
    out["row_id"] = [f"S_{sheet_name}_{i}" for i in range(len(out))]

    out = out[out["amount_signed"].notna()]
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)
    return out, {"sheet": sheet_name, "inv_col": inv_col, "date_col": date_col, "amt_col": amt_col, "desc_col": desc_col}

def combine_supplier_workbook(uploaded_file, invoice_hint: str, date_hint: str, amount_hint: str, desc_hint: str):
    xls = pd.ExcelFile(uploaded_file)
    all_norm = []
    audit = []

    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sheet)
        if df is None:
            audit.append({"sheet": sheet, "kept": False, "reason": "no table"})
            continue

        if meta["rows"] < 5 or meta["cols"] < 3:
            audit.append({"sheet": sheet, "kept": False, "reason": "too small"})
            continue

        norm, used = normalize_supplier_sheet(df, sheet, invoice_hint, date_hint, amount_hint, desc_hint)
        if norm.empty:
            audit.append({"sheet": sheet, "kept": False, "reason": "normalized empty"})
            continue

        audit.append({"sheet": sheet, "kept": True, "reason": "ok", **used, "rows": len(norm)})
        all_norm.append(norm)

    audit_df = pd.DataFrame(audit)
    if not all_norm:
        return pd.DataFrame(), audit_df

    combined = pd.concat(all_norm, ignore_index=True)

    combined["dedupe_key"] = (
        combined["invoice_key"].fillna("") + "|" +
        combined["doc_date"].dt.strftime("%Y-%m-%d").fillna("") + "|" +
        combined["amt_r2"].fillna(0).astype(str)
    )

    combined = combined.sort_values(["doc_date"]).drop_duplicates("dedupe_key", keep="first").reset_index(drop=True)
    return combined, audit_df

def normalize_ledger(df: pd.DataFrame, external_hint: str, date_hint: str, amount_hint: str, desc_hint: str, doc_type_hint: str, supplier_invoice_set: set, flip_sign: bool):
    ext_col = pick_col_by_hint(df, external_hint) or infer_col_by_type(df, "external_doc")
    date_col = pick_col_by_hint(df, date_hint) or infer_col_by_type(df, "date")
    amt_col = pick_col_by_hint(df, amount_hint) or infer_col_by_type(df, "amount")
    desc_col = pick_col_by_hint(df, desc_hint) or infer_col_by_type(df, "description")
    doc_type_col = pick_col_by_hint(df, doc_type_hint)

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=True)) if date_col in df.columns else pd.NaT
    out["external_doc_raw"] = df[ext_col].map(to_str) if ext_col in df.columns else ""
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""
    out["doc_type"] = df[doc_type_col].map(to_str) if doc_type_col and doc_type_col in df.columns else ""

    amt = df[amt_col].map(to_num) if amt_col in df.columns else np.nan
    out["amount_signed"] = (-amt) if flip_sign else amt

    out["amt_r2"] = out["amount_signed"].map(round2)
    out["amt_r0"] = out["amount_signed"].map(round0)

    out["invoice_key_extracted"] = out["external_doc_raw"].apply(lambda x: ledger_invoice_key(x, supplier_invoice_set))
    out["txn_type"] = [
        classify_ledger_txn(dt, ext, ds)
        for dt, ext, ds in zip(out["doc_type"].tolist(), out["external_doc_raw"].tolist(), out["description"].tolist())
    ]

    out["docid"] = out["external_doc_raw"].apply(extract_docid)

    out["row_id"] = [f"L_{i}" for i in range(len(out))]

    out = out[out["amount_signed"].notna()]
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)

    used = {"ext_col": ext_col, "date_col": date_col, "amt_col": amt_col, "desc_col": desc_col, "doc_type_col": doc_type_col}
    return out, used


def normalize_statement_like_supplier(df: pd.DataFrame, sheet_name: str, ref_hint: str, date_hint: str, debit_hint: str, credit_hint: str, desc_hint: str, balance_hint: str):
    ref_col = pick_col_by_hint(df, ref_hint) or pick_col_by_hint(df, "reference") or pick_col_by_hint(df, "ref") or infer_col_by_type(df, "invoice")
    date_col = pick_col_by_hint(df, date_hint) or infer_col_by_type(df, "date")
    debit_col = pick_col_by_hint(df, debit_hint) or pick_col_by_hint(df, "debit")
    credit_col = pick_col_by_hint(df, credit_hint) or pick_col_by_hint(df, "credit")
    desc_col = pick_col_by_hint(df, desc_hint) or infer_col_by_type(df, "description")
    bal_col = pick_col_by_hint(df, balance_hint) or pick_col_by_hint(df, "balance")

    out = pd.DataFrame()
    out["doc_date"] = df[date_col].map(lambda v: to_date(v, dayfirst=False)) if date_col in df.columns else pd.NaT
    out["reference_raw"] = df[ref_col].map(to_str) if ref_col in df.columns else ""
    out["description"] = df[desc_col].map(to_str) if desc_col in df.columns else ""

    debit = df[debit_col].map(to_num) if debit_col and debit_col in df.columns else np.nan
    credit = df[credit_col].map(to_num) if credit_col and credit_col in df.columns else np.nan

    debit = debit.fillna(0) if isinstance(debit, pd.Series) else pd.Series([0] * len(df))
    credit = credit.fillna(0) if isinstance(credit, pd.Series) else pd.Series([0] * len(df))

    out["amount_signed"] = debit - credit
    out["abs_amount"] = out["amount_signed"].abs()
    out["docid"] = out["reference_raw"].apply(extract_docid)

    if bal_col and bal_col in df.columns:
        out["balance"] = df[bal_col].map(to_num)
    else:
        out["balance"] = np.nan

    out["sheet_name"] = sheet_name
    out["row_id"] = [f"ST_{sheet_name}_{i}" for i in range(len(out))]

    out = out.dropna(how="all")
    out = out.dropna(subset=["doc_date"]).reset_index(drop=True)

    used = {
        "sheet": sheet_name,
        "ref_col": ref_col,
        "date_col": date_col,
        "debit_col": debit_col,
        "credit_col": credit_col,
        "desc_col": desc_col,
        "balance_col": bal_col,
    }
    return out, used

def combine_statement_workbook(uploaded_file, ref_hint: str, date_hint: str, debit_hint: str, credit_hint: str, desc_hint: str, balance_hint: str):
    xls = pd.ExcelFile(uploaded_file)
    all_norm = []
    audit = []

    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sheet)
        if df is None:
            audit.append({"sheet": sheet, "kept": False, "reason": "no table"})
            continue

        if meta["rows"] < 5 or meta["cols"] < 3:
            audit.append({"sheet": sheet, "kept": False, "reason": "too small"})
            continue

        norm, used = normalize_statement_like_supplier(df, sheet, ref_hint, date_hint, debit_hint, credit_hint, desc_hint, balance_hint)
        if norm.empty:
            audit.append({"sheet": sheet, "kept": False, "reason": "normalized empty"})
            continue

        audit.append({"sheet": sheet, "kept": True, "reason": "ok", **used, "rows": len(norm)})
        all_norm.append(norm)

    audit_df = pd.DataFrame(audit)
    if not all_norm:
        return pd.DataFrame(), audit_df

    combined = pd.concat(all_norm, ignore_index=True)

    combined = combined.dropna(subset=["doc_date"]).reset_index(drop=True)
    return combined, audit_df


def reconcile_docid(statement_df: pd.DataFrame, ledger_df: pd.DataFrame, amount_tol: float):
    stx = statement_df.copy()
    ltx = ledger_df.copy()

    stx = stx[stx["docid"].fillna("").astype(str) != ""].copy()
    ltx = ltx[ltx["docid"].fillna("").astype(str) != ""].copy()

    s_group = (
        stx.groupby("docid", as_index=False)
        .agg(
            doc_date=("doc_date", "min"),
            ref=("reference_raw", "first"),
            details=("description", "first"),
            amount=("abs_amount", "sum"),
        )
    )

    l_group = (
        ltx.groupby("docid", as_index=False)
        .agg(
            doc_date=("doc_date", "min"),
            ref=("external_doc_raw", "first"),
            details=("description", "first"),
            amount=("amount_signed", lambda s: float(pd.Series(s).abs().sum())),
        )
    )
    l_group["amount"] = l_group["amount"].abs()

    merged = pd.merge(
        s_group.add_prefix("Stmt_"),
        l_group.add_prefix("Ledg_"),
        left_on="Stmt_docid",
        right_on="Ledg_docid",
        how="outer",
        indicator=True
    )
    merged["key"] = merged["Stmt_docid"].combine_first(merged["Ledg_docid"])
    merged["variance"] = merged["Stmt_amount"] - merged["Ledg_amount"]
    merged["status"] = merged.apply(
        lambda r: (
            "Post in ledger" if r["_merge"] == "left_only"
            else "Include on statement" if r["_merge"] == "right_only"
            else "Matched" if abs(float(r["variance"])) <= amount_tol
            else "Amount mismatch"
        ),
        axis=1
    )

    stmt_only = merged[merged["_merge"] == "left_only"].copy()
    ledg_only = merged[merged["_merge"] == "right_only"].copy()
    mismatches = merged[(merged["_merge"] == "both") & (merged["status"] == "Amount mismatch")].copy()
    matched = merged[(merged["_merge"] == "both") & (merged["status"] == "Matched")].copy()

    left_table = pd.DataFrame({
        "Date": ledg_only["Ledg_doc_date"],
        "Ref": ledg_only["Ledg_docid"],
        "Details": ledg_only["Ledg_details"],
        "Amount": ledg_only["Ledg_amount"],
        "Action": "Include on statement"
    })

    right_table = pd.DataFrame({
        "Date": stmt_only["Stmt_doc_date"],
        "Ref": stmt_only["Stmt_docid"],
        "Details": stmt_only["Stmt_details"],
        "Amount": stmt_only["Stmt_amount"],
        "Action": "Post in ledger"
    })

    match_detail = pd.DataFrame({
        "match_method": "docid_total",
        "status": merged["status"],
        "match_key": merged["key"],
        "statement_date": merged["Stmt_doc_date"],
        "statement_ref": merged["Stmt_docid"],
        "statement_details": merged["Stmt_details"],
        "statement_amount": merged["Stmt_amount"],
        "ledger_date": merged["Ledg_doc_date"],
        "ledger_ref": merged["Ledg_docid"],
        "ledger_details": merged["Ledg_details"],
        "ledger_amount": merged["Ledg_amount"],
        "difference_statement_minus_ledger": merged["variance"],
    })

    return {
        "match_detail": match_detail,
        "left_table": left_table.sort_values(["Date", "Ref"], na_position="last"),
        "right_table": right_table.sort_values(["Date", "Ref"], na_position="last"),
        "mismatches": mismatches,
        "matched": matched,
        "stmt_only": stmt_only,
        "ledg_only": ledg_only,
    }


def mk_detail(srow, lrow, status, method, match_key):
    return {
        "match_method": method,
        "status": status,
        "match_key": match_key,
        "supplier_date": None if srow is None else srow.get("doc_date"),
        "supplier_ref": "" if srow is None else srow.get("ref"),
        "supplier_details": "" if srow is None else srow.get("details"),
        "supplier_amount": np.nan if srow is None else srow.get("amount"),
        "ledger_date": None if lrow is None else lrow.get("doc_date"),
        "ledger_ref": "" if lrow is None else lrow.get("ref"),
        "ledger_details": "" if lrow is None else lrow.get("details"),
        "ledger_amount": np.nan if lrow is None else lrow.get("amount"),
        "difference_supplier_minus_ledger": np.nan if (srow is None or lrow is None) else float(srow.get("amount") - lrow.get("amount")),
    }

def reconcile_invoice_style(supplier: pd.DataFrame, ledger: pd.DataFrame, amount_tol: float, date_window_days: int):
    details = []

    s_inv = supplier[supplier["invoice_key"].fillna("") != ""].copy()
    l_inv = ledger[ledger["invoice_key_extracted"].fillna("") != ""].copy()

    s_agg = (
        s_inv.groupby("invoice_key", as_index=False)
        .agg(doc_date=("doc_date", "min"),
             ref=("invoice_no_raw", "first"),
             details=("description", "first"),
             amount=("amount_signed", "sum"))
    )

    l_agg = (
        l_inv.groupby("invoice_key_extracted", as_index=False)
        .agg(doc_date=("doc_date", "min"),
             ref=("external_doc_raw", "first"),
             details=("description", "first"),
             amount=("amount_signed", "sum"))
        .rename(columns={"invoice_key_extracted": "invoice_key"})
    )

    s_keys = set(s_agg["invoice_key"].unique().tolist())
    l_keys = set(l_agg["invoice_key"].unique().tolist())
    all_keys = sorted(list(s_keys.union(l_keys)))

    l_map = {r["invoice_key"]: r for _, r in l_agg.iterrows()}
    s_map = {r["invoice_key"]: r for _, r in s_agg.iterrows()}

    matched_supplier_keys = set()
    matched_ledger_keys = set()

    for k in all_keys:
        srow = s_map.get(k)
        lrow = l_map.get(k)

        if srow is None and lrow is not None:
            details.append(mk_detail(None, lrow, "Missing on Supplier", "invoice_key_total", k))
            continue
        if lrow is None and srow is not None:
            details.append(mk_detail(srow, None, "Missing in Ledger", "invoice_key_total", k))
            continue

        diff = float(srow["amount"] - lrow["amount"])
        status = "Matched" if abs(diff) <= amount_tol else "Amount mismatch"
        details.append(mk_detail(srow, lrow, status, "invoice_key_total", k))
        matched_supplier_keys.add(k)
        matched_ledger_keys.add(k)

    s_left = supplier[~supplier["invoice_key"].isin(matched_supplier_keys)].copy()
    l_left = ledger[~ledger["invoice_key_extracted"].isin(matched_ledger_keys)].copy()

    s_pay = s_left[s_left["invoice_key"].fillna("") == ""].copy()
    l_pay = l_left[l_left["txn_type"] == "payment"].copy()

    used_l = set()
    l_groups = l_pay.groupby("amt_r0")

    for _, sr in s_pay.iterrows():
        amt_key = sr["amt_r0"]
        if pd.isna(amt_key) or amt_key not in l_groups.groups:
            continue

        candidates = l_groups.get_group(amt_key).copy()
        candidates = candidates[~candidates["row_id"].isin(used_l)]
        if candidates.empty:
            continue

        candidates["date_diff"] = candidates["doc_date"].apply(lambda d: date_diff_days(d, sr["doc_date"]))
        candidates = candidates[candidates["date_diff"] <= date_window_days]
        if candidates.empty:
            continue

        best = candidates.sort_values(["date_diff"]).iloc[0]

        srow = {"doc_date": sr["doc_date"], "ref": sr["invoice_no_raw"], "details": sr["description"], "amount": sr["amount_signed"]}
        lrow = {"doc_date": best["doc_date"], "ref": best["external_doc_raw"], "details": best["description"], "amount": best["amount_signed"]}
        diff = float(srow["amount"] - lrow["amount"])
        status = "Matched" if abs(diff) <= amount_tol else "Amount mismatch"
        details.append(mk_detail(srow, lrow, status, "payment_amount_date_whole", f"PAY|{amt_key}"))
        used_l.add(best["row_id"])

    match_detail = pd.DataFrame(details)

    missing_in_ledger = match_detail[match_detail["status"] == "Missing in Ledger"].copy()
    missing_on_supplier = match_detail[match_detail["status"] == "Missing on Supplier"].copy()
    amount_mismatch = match_detail[match_detail["status"] == "Amount mismatch"].copy()

    left_table = pd.DataFrame({
        "Date": missing_on_supplier["ledger_date"],
        "Ref": missing_on_supplier["ledger_ref"],
        "Details": missing_on_supplier["ledger_details"],
        "Amount": missing_on_supplier["ledger_amount"],
        "Action": ""
    })

    right_table = pd.DataFrame({
        "Date": missing_in_ledger["supplier_date"],
        "Ref": missing_in_ledger["supplier_ref"],
        "Details": missing_in_ledger["supplier_details"],
        "Amount": missing_in_ledger["supplier_amount"],
        "Action": ""
    })

    return {
        "match_detail": match_detail,
        "left_table": left_table,
        "right_table": right_table,
        "missing_in_ledger": missing_in_ledger,
        "missing_on_supplier": missing_on_supplier,
        "amount_mismatch": amount_mismatch,
    }


def parse_cell(addr: str):
    addr = (addr or "").strip().upper()
    m = re.match(r"^([A-Z]+)(\d+)$", addr)
    if not m:
        raise ValueError(f"Invalid cell address: {addr}. Use format like B16.")
    col_letters, row_str = m.group(1), m.group(2)
    return int(row_str), column_index_from_string(col_letters)

def excel_safe(v):
    if v is None:
        return None
    if pd.isna(v):
        return None
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.datetime64):
        dtv = pd.to_datetime(v, errors="coerce")
        if pd.isna(dtv):
            return None
        return dtv.to_pydatetime()
    if isinstance(v, date) and not isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    return v

def clear_range(ws, start_row, start_col, n_rows, n_cols):
    for r in range(start_row, start_row + n_rows):
        for c in range(start_col, start_col + n_cols):
            ws.cell(r, c).value = None

def write_table(ws, start_row, start_col, df: pd.DataFrame, action_col: bool, max_rows=6000):
    if df is None or df.empty:
        return
    cols = ["Date", "Ref", "Details", "Amount"]
    if action_col:
        cols.append("Action")
    df2 = df.copy()
    if "Action" not in df2.columns:
        df2["Action"] = ""
    df2 = df2[cols]
    rows = min(len(df2), max_rows)
    for i in range(rows):
        for j, c in enumerate(cols):
            ws.cell(start_row + i, start_col + j).value = excel_safe(df2.iloc[i, j])

def write_df_full(ws, df: pd.DataFrame, max_rows=80000):
    if df is None:
        return
    ws.append(list(df.columns))
    rows = min(len(df), max_rows)
    for i in range(rows):
        ws.append([excel_safe(v) for v in list(df.iloc[i].values)])
    for c in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

def find_row_by_label(ws, label, label_col=2):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, label_col).value
        if isinstance(v, str) and v.strip() == label:
            return r
    return None

def export_pack_recon_format(template_upload, left_df, right_df, stmt_balance, ledg_balance, as_at_dt, supplier_name, start_row=9, totals_row=29):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    template_upload.seek(0)
    with open(tmp_path, "wb") as f:
        f.write(template_upload.read())

    wb = load_workbook(tmp_path)
    ws = wb[wb.sheetnames[0]]

    ws["F2"].value = supplier_name
    ws["K5"].value = as_at_dt
    ws["K5"].number_format = "dd/mm/yyyy"

    needed_rows = max(len(left_df), len(right_df), 1)
    available_rows = totals_row - start_row
    insert_n = max(0, needed_rows - available_rows)

    if insert_n > 0:
        ws.insert_rows(totals_row, amount=insert_n)
        totals_row += insert_n

    template_style_row = start_row
    for r in range(start_row + available_rows, start_row + needed_rows):
        for c in range(2, 13):
            src = ws.cell(template_style_row, c)
            dst = ws.cell(r, c)
            dst._style = pycopy(src._style)
            dst.number_format = src.number_format

    for r in range(start_row, start_row + needed_rows):
        for c in range(2, 13):
            ws.cell(r, c).value = None

    LEFT = {"Date": 2, "Ref": 3, "Details": 4, "Amount": 5, "Action": 6}
    RIGHT = {"Date": 8, "Ref": 9, "Details": 10, "Amount": 11, "Action": 12}
    date_fmt = "dd/mm/yy"

    for i in range(needed_rows):
        r = start_row + i

        if i < len(left_df):
            it = left_df.iloc[i].to_dict()
            ws.cell(r, LEFT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, LEFT["Date"]).number_format = date_fmt
            ws.cell(r, LEFT["Ref"]).value = it.get("Ref")
            ws.cell(r, LEFT["Details"]).value = it.get("Details")
            ws.cell(r, LEFT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, LEFT["Action"]).value = it.get("Action", "")

        if i < len(right_df):
            it = right_df.iloc[i].to_dict()
            ws.cell(r, RIGHT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, RIGHT["Date"]).number_format = date_fmt
            ws.cell(r, RIGHT["Ref"]).value = it.get("Ref")
            ws.cell(r, RIGHT["Details"]).value = it.get("Details")
            ws.cell(r, RIGHT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, RIGHT["Action"]).value = it.get("Action", "")

    left_sum_range = f"E{start_row}:E{start_row + needed_rows - 1}"
    right_sum_range = f"K{start_row}:K{start_row + needed_rows - 1}"
    ws[f"E{totals_row}"].value = f"=SUM({left_sum_range})"
    ws[f"K{totals_row}"].value = f"=SUM({right_sum_range})"

    r_stmt = find_row_by_label(ws, "Balance as per Supplier Statement")
    r_ledg = find_row_by_label(ws, "Balance as per Creditors Ledger")
    if r_stmt:
        ws.cell(r_stmt, 10).value = excel_safe(stmt_balance)
    if r_ledg:
        ws.cell(r_ledg, 10).value = excel_safe(ledg_balance)

    r_adj_sup = find_row_by_label(ws, "Add: Adjustments to be made by Supplier")
    r_adj_books = find_row_by_label(ws, "Add: Adjustments to be made in our Books")
    if r_adj_sup:
        ws.cell(r_adj_sup, 10).value = f"=+E{totals_row}"
    if r_adj_books:
        ws.cell(r_adj_books, 10).value = f"=+K{totals_row}"

    wb.save(tmp_path)
    return tmp_path

def export_pack_generic(template_upload, results: dict, left_start_cell: str, right_start_cell: str, action_col: bool):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name

    template_upload.seek(0)
    with open(tmp_path, "wb") as f:
        f.write(template_upload.read())

    wb = load_workbook(tmp_path)
    ws = wb[wb.sheetnames[0]]

    lrow, lcol = parse_cell(left_start_cell)
    rrow, rcol = parse_cell(right_start_cell)

    clear_range(ws, lrow, lcol, 7000, 5 if action_col else 4)
    clear_range(ws, rrow, rcol, 7000, 5 if action_col else 4)

    write_table(ws, lrow, lcol, results["left_table"], action_col=action_col)
    write_table(ws, rrow, rcol, results["right_table"], action_col=action_col)

    for name in ["Match_Detail", "missing_in_ledger", "missing_on_supplier", "amount_mismatch"]:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    md = wb.create_sheet("Match_Detail")
    write_df_full(md, results.get("match_detail"))

    mi = wb.create_sheet("missing_in_ledger")
    write_df_full(mi, results.get("missing_in_ledger"))

    ms = wb.create_sheet("missing_on_supplier")
    write_df_full(ms, results.get("missing_on_supplier"))

    am = wb.create_sheet("amount_mismatch")
    write_df_full(am, results.get("amount_mismatch"))

    wb.save(tmp_path)
    return tmp_path


st.title("Lets Reconcile")

with st.sidebar:
    supplier_file = st.file_uploader("Supplier workbook", type=["xlsx", "xls"])
    ledger_file = st.file_uploader("Ledger workbook", type=["xlsx", "xls"])
    template_file = st.file_uploader("Recon template", type=["xlsx", "xls"])

    st.subheader("Mode")
    recon_mode = st.selectbox(
        "Reconciliation type",
        [
            "Supplier Statement (Debit/Credit + HREINV/HRECRN) vs Ledger",
            "Invoice workbook (invoice key AE...) vs Ledger"
        ],
        index=0
    )

    st.subheader("Supplier Statement hints")
    stmt_ref_hint = st.text_input("Statement reference hint", value="reference")
    stmt_date_hint = st.text_input("Statement date hint", value="date")
    stmt_debit_hint = st.text_input("Statement debit hint", value="debit")
    stmt_credit_hint = st.text_input("Statement credit hint", value="credit")
    stmt_desc_hint = st.text_input("Statement description hint", value="description")
    stmt_balance_hint = st.text_input("Statement balance hint", value="balance")

    st.subheader("Invoice workbook hints")
    st.caption("Only used in Invoice workbook mode")
    supplier_invoice_hint = st.text_input("Supplier invoice hint", value="invoice")
    supplier_date_hint = st.text_input("Supplier date hint", value="date")
    supplier_amount_hint = st.text_input("Supplier amount hint", value="amount")
    supplier_desc_hint = st.text_input("Supplier description hint", value="desc")

    st.subheader("Ledger hints")
    ledger_external_hint = st.text_input("Ledger external doc hint", value="external")
    ledger_date_hint = st.text_input("Ledger posting date hint", value="posting")
    ledger_amount_hint = st.text_input("Ledger amount hint", value="amount")
    ledger_desc_hint = st.text_input("Ledger description hint", value="description")
    ledger_doc_type_hint = st.text_input("Ledger doc type hint", value="type")

    st.subheader("Settings")
    flip_ledger_sign = st.checkbox("Flip ledger sign", value=True)
    amount_tolerance = st.number_input("Amount tolerance", min_value=0.0, value=0.05, step=0.01)
    date_window_days = st.number_input("Date window days", min_value=0, value=14, step=1)

    st.subheader("Export")
    use_recon_format_layout = st.checkbox("Template is RECON FORMART layout", value=True)
    supplier_name = st.text_input("Supplier name (for template header)", value="SUPPLIER")

    st.caption("Only used when RECON FORMART layout is off")
    left_start_cell = st.text_input("Left table start cell", value="B16")
    right_start_cell = st.text_input("Right table start cell", value="H16")
    template_has_action = st.checkbox("Template has Action column", value=True)

run_btn = st.button("Run reconciliation", type="primary", use_container_width=True)

if run_btn:
    if not (supplier_file and ledger_file and template_file):
        st.error("Upload supplier, ledger, and template files.")
        st.stop()

    with st.spinner("Loading ledger workbook..."):
        xls = pd.ExcelFile(ledger_file)
        best_df, meta = None, None
        for sheet in xls.sheet_names:
            raw = pd.read_excel(xls, sheet_name=sheet, header=None, dtype=object)
            df, m = detect_best_table_in_sheet(raw, sheet_name=sheet)
            if df is None:
                continue
            if best_df is None or m["score"] > meta["score"]:
                best_df, meta = df, m

    if best_df is None:
        st.error("No usable ledger table detected.")
        st.stop()

    st.subheader("Ledger detected table preview")
    st.write(f"Sheet: {meta['sheet_name']}  Header row: {meta['header_row'] + 1}")
    st.dataframe(best_df.head(50), use_container_width=True)

    if recon_mode.startswith("Supplier Statement"):
        with st.spinner("Combining statement sheets and extracting HREINV/HRECRN..."):
            stmt_combined, stmt_audit = combine_statement_workbook(
                supplier_file,
                ref_hint=stmt_ref_hint,
                date_hint=stmt_date_hint,
                debit_hint=stmt_debit_hint,
                credit_hint=stmt_credit_hint,
                desc_hint=stmt_desc_hint,
                balance_hint=stmt_balance_hint
            )

        st.subheader("Statement sheet audit")
        st.dataframe(stmt_audit, use_container_width=True)

        if stmt_combined.empty:
            st.error("No usable statement table detected.")
            st.stop()

        st.subheader("Statement combined preview")
        st.dataframe(stmt_combined.head(80), use_container_width=True)

        supplier_invoice_set = set()

        with st.spinner("Normalizing ledger and extracting HREINV/HRECRN..."):
            ledger_norm, ledger_used = normalize_ledger(
                best_df,
                external_hint=ledger_external_hint,
                date_hint=ledger_date_hint,
                amount_hint=ledger_amount_hint,
                desc_hint=ledger_desc_hint,
                doc_type_hint=ledger_doc_type_hint,
                supplier_invoice_set=supplier_invoice_set,
                flip_sign=flip_ledger_sign
            )

        st.write("Ledger columns used")
        st.write(ledger_used)

        st.subheader("Ledger normalized preview")
        st.dataframe(ledger_norm.head(80), use_container_width=True)

        with st.spinner("Reconciling by HREINV/HRECRN totals..."):
            results_docid = reconcile_docid(
                stmt_combined,
                ledger_norm,
                amount_tol=float(amount_tolerance)
            )

        st.subheader("Recon preview")
        c1, c2 = st.columns(2)
        with c1:
            st.write("Left table: ledger missing on statement")
            st.dataframe(results_docid["left_table"].head(500), use_container_width=True)
        with c2:
            st.write("Right table: statement missing in ledger")
            st.dataframe(results_docid["right_table"].head(500), use_container_width=True)

        st.subheader("Match detail")
        st.dataframe(results_docid["match_detail"].head(600), use_container_width=True)

        stmt_balance = float(stmt_combined["balance"].dropna().iloc[-1]) if stmt_combined["balance"].dropna().shape[0] else np.nan
        ledg_bal_col = pick_col_by_hint(best_df, "balance") or pick_col_by_hint(best_df, "balance (lcy)") or ""
        ledg_balance = float(best_df[ledg_bal_col].map(to_num).dropna().iloc[-1]) if ledg_bal_col and best_df[ledg_bal_col].map(to_num).dropna().shape[0] else np.nan

        stmt_max_date = pd.to_datetime(stmt_combined["doc_date"], errors="coerce").max()
        ledg_max_date = pd.to_datetime(ledger_norm["doc_date"], errors="coerce").max()
        cands = [d for d in [stmt_max_date, ledg_max_date] if pd.notna(d)]
        as_at = max(cands).to_pydatetime() if cands else datetime.now()

        with st.spinner("Writing output workbook..."):
            if use_recon_format_layout:
                out_path = export_pack_recon_format(
                    template_upload=template_file,
                    left_df=results_docid["left_table"],
                    right_df=results_docid["right_table"],
                    stmt_balance=stmt_balance,
                    ledg_balance=ledg_balance,
                    as_at_dt=as_at,
                    supplier_name=supplier_name
                )
                wb = load_workbook(out_path)
                for sh in ["Mismatches", "Summary"]:
                    if sh in wb.sheetnames:
                        wb.remove(wb[sh])

                ws_m = wb.create_sheet("Mismatches")
                ws_m.append(["Ref", "Statement Amount", "Ledger Amount", "Variance (Stmt-Ledg)", "Statement Date", "Ledger Date", "Statement Details", "Ledger Details"])
                md = results_docid["match_detail"]
                mm = md[md["status"] == "Amount mismatch"].copy()
                for _, r in mm.iterrows():
                    ws_m.append([
                        r["match_key"],
                        excel_safe(r["statement_amount"]),
                        excel_safe(r["ledger_amount"]),
                        excel_safe(r["difference_statement_minus_ledger"]),
                        excel_safe(r["statement_date"]),
                        excel_safe(r["ledger_date"]),
                        r["statement_details"],
                        r["ledger_details"],
                    ])

                ws_s = wb.create_sheet("Summary")
                ws_s.append(["Metric", "Value"])
                ws_s.append(["As at", as_at.date().isoformat()])
                ws_s.append(["Statement balance", excel_safe(stmt_balance)])
                ws_s.append(["Ledger balance", excel_safe(ledg_balance)])
                ws_s.append(["Statement-only refs (post in ledger)", int(len(results_docid["right_table"]))])
                ws_s.append(["Ledger-only refs (include on statement)", int(len(results_docid["left_table"]))])
                ws_s.append(["Amount mismatches (investigate)", int(len(mm))])
                ws_s.append(["Matched refs", int((md["status"] == "Matched").sum())])

                wb.save(out_path)
            else:
                generic_pack = {
                    "left_table": results_docid["left_table"],
                    "right_table": results_docid["right_table"],
                    "match_detail": results_docid["match_detail"],
                    "missing_in_ledger": results_docid["stmt_only"],
                    "missing_on_supplier": results_docid["ledg_only"],
                    "amount_mismatch": results_docid["mismatches"],
                }
                out_path = export_pack_generic(
                    template_upload=template_file,
                    results=generic_pack,
                    left_start_cell=left_start_cell,
                    right_start_cell=right_start_cell,
                    action_col=template_has_action,
                )

        with open(out_path, "rb") as f:
            st.download_button(
                "Download recon output",
                data=f,
                file_name=f"recon_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    else:
        with st.spinner("Combining supplier sheets..."):
            supplier_combined, supplier_audit = combine_supplier_workbook(
                supplier_file,
                invoice_hint=supplier_invoice_hint,
                date_hint=supplier_date_hint,
                amount_hint=supplier_amount_hint,
                desc_hint=supplier_desc_hint
            )

        st.subheader("Supplier sheet audit")
        st.dataframe(supplier_audit, use_container_width=True)

        if supplier_combined.empty:
            st.error("No usable supplier transaction table detected across sheets.")
            st.stop()

        blank_invoice_pct = float((supplier_combined["invoice_key"].fillna("") == "").mean())
        st.write(f"Supplier rows: {len(supplier_combined):,}")
        st.write(f"Supplier invoice_key blank rate: {blank_invoice_pct:.1%}")

        supplier_invoice_set = set(supplier_combined["invoice_key"].dropna().astype(str).unique().tolist())
        supplier_invoice_set = {x for x in supplier_invoice_set if x}

        st.subheader("Supplier combined preview")
        st.dataframe(supplier_combined.head(50), use_container_width=True)

        with st.spinner("Normalizing ledger and extracting invoice keys..."):
            ledger_norm, ledger_used = normalize_ledger(
                best_df,
                external_hint=ledger_external_hint,
                date_hint=ledger_date_hint,
                amount_hint=ledger_amount_hint,
                desc_hint=ledger_desc_hint,
                doc_type_hint=ledger_doc_type_hint,
                supplier_invoice_set=supplier_invoice_set,
                flip_sign=flip_ledger_sign
            )

        st.write("Ledger columns used")
        st.write(ledger_used)

        extracted_blank_pct = float((ledger_norm["invoice_key_extracted"].fillna("") == "").mean())
        st.write(f"Ledger rows: {len(ledger_norm):,}")
        st.write(f"Ledger invoice_key extracted blank rate: {extracted_blank_pct:.1%}")

        st.subheader("Ledger normalized preview")
        st.dataframe(ledger_norm.head(50), use_container_width=True)

        with st.spinner("Reconciling..."):
            results = reconcile_invoice_style(
                supplier_combined,
                ledger_norm,
                amount_tol=float(amount_tolerance),
                date_window_days=int(date_window_days),
            )

        st.subheader("Recon preview")
        c1, c2 = st.columns(2)
        with c1:
            st.write("Left table: ledger missing on supplier")
            st.dataframe(results["left_table"].head(200), use_container_width=True)
        with c2:
            st.write("Right table: supplier missing in ledger")
            st.dataframe(results["right_table"].head(200), use_container_width=True)

        st.subheader("Match detail")
        st.dataframe(results["match_detail"].head(400), use_container_width=True)

        with st.spinner("Writing output workbook..."):
            out_path = export_pack_generic(
                template_upload=template_file,
                results=results,
                left_start_cell=left_start_cell,
                right_start_cell=right_start_cell,
                action_col=template_has_action,
            )

        with open(out_path, "rb") as f:
            st.download_button(
                "Download recon output",
                data=f,
                file_name=f"recon_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
